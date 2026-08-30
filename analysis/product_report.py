"""Generacion del informe Excel multi-pestana de un analisis de producto: una pestana
principal de sintesis (Analisis_Activo) mas las 4 pestanas de datos crudos que la alimentaron
(Conflictos_Geopoliticos, Bancos_Centrales_Tipos, Cadenas_Suministro_Energia,
Sentimiento_Fondos), para que el cruce de variables del LLM sea auditable -se puede verificar
exactamente que datos vio- en vez de una caja negra.
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analysis.schemas import ProductForecast
from config import BASE_DIR
from utils.logging_conf import get_logger

logger = get_logger(__name__)

REPORTS_DIR = BASE_DIR / "data" / "product_reports"

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="E2E8F0", bold=True)
TITLE_FONT = Font(bold=True, size=14)
LABEL_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def _slug(text: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", text.strip().lower()).strip("-")
    return slug[:40] or "producto"


def _write_table(ws: Worksheet, headers: list[str], rows: list[list], start_row: int = 1) -> int:
    """Escribe una tabla con cabecera estilizada. Devuelve la fila siguiente a la ultima escrita."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30
    return start_row + len(rows) + 2


def _write_main_sheet(ws: Worksheet, producto: str, forecast: ProductForecast) -> None:
    ws["A1"] = f"Tesis de inversion: {producto}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generado: {datetime.now().isoformat(timespec='seconds')}"

    hyp = forecast.hipotesis_falsable
    rows = [
        ("Direccion", forecast.direccion),
        ("Probabilidad", f"{forecast.probabilidad:.0%}"),
        ("Horizonte (meses)", forecast.horizonte_meses),
        ("Nivel de confianza", f"{forecast.nivel_confianza:.0%}"),
        ("Conflictos considerados", ", ".join(forecast.conflictos_considerados) or "Ninguno directamente relevante"),
        ("", ""),
        ("TESIS DE INVERSION (cruce de los 4 bloques)", ""),
        ("Tesis de inversion", forecast.tesis_inversion),
        ("", ""),
        ("SINTESIS POR BLOQUE", ""),
        ("Bloque 2 - Bancos centrales", forecast.resumen_bancos_centrales),
        ("Bloque 3 - Energia / suministro", forecast.resumen_energia_suministro),
        ("Bloque 4 - Sentimiento", forecast.resumen_sentimiento),
        ("", ""),
        ("Escenario alternativo", forecast.escenario_alternativo),
        ("Limitaciones declaradas", forecast.limitaciones),
        ("", ""),
        ("HIPOTESIS FALSABLE (track record separado del predictor de producto)", ""),
        ("Enunciado", hyp.enunciado),
        ("Ticker de validacion", hyp.ticker_validacion),
        ("Comparador", hyp.comparador),
        ("Valor umbral", hyp.valor_umbral),
        ("Valor base al emitir", hyp.valor_base_al_emitir),
        ("Fecha limite de revision", str(hyp.fecha_limite_revision)),
        ("Descripcion de la metrica", hyp.descripcion_metrica),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = LABEL_FONT
        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.alignment = WRAP

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 95


def build_product_workbook(
    producto: str,
    conflicts: list[dict],
    central_banks: dict,
    energy: list[dict],
    sentiment: list[dict],
    forecast: ProductForecast,
) -> Path:
    """Construye el workbook de 5 pestanas y lo guarda en `data/product_reports/`. Devuelve la
    ruta al archivo generado."""
    wb = Workbook()

    ws_main = wb.active
    ws_main.title = "Analisis_Activo"
    _write_main_sheet(ws_main, producto, forecast)

    ws_conflicts = wb.create_sheet("Conflictos_Geopoliticos")
    considerados = set(forecast.conflictos_considerados)
    _write_table(
        ws_conflicts,
        ["Nombre", "Considerado en la tesis", "Continente", "Pais principal",
         "Latitud", "Longitud", "Inicio aprox.", "Paises afectados"],
        [[
            c["nombre"], "SI" if c["nombre"] in considerados else "",
            c["continente"], c["pais_principal"], c["latitud"], c["longitud"],
            c["inicio_aproximado"], ", ".join(c["paises"]),
        ] for c in conflicts],
    )

    ws_banks = wb.create_sheet("Bancos_Centrales_Tipos")
    ws_banks.cell(row=1, column=1, value="Titulares oficiales recientes").font = LABEL_FONT
    next_row = _write_table(
        ws_banks,
        ["Banco central", "Titular", "Publicado", "URL"],
        [[h["banco"], h["titular"], h["publicado"], h["url"]] for h in central_banks["titulares"]],
        start_row=2,
    )
    ws_banks.cell(row=next_row, column=1, value="Proxies de mercado de tipos de interes").font = LABEL_FONT
    _write_table(
        ws_banks,
        ["Ticker", "Nombre", "Precio actual", "Variacion (%)"],
        [[p["ticker"], p["nombre"], p["precio_actual"], p["variacion_pct"]] for p in central_banks["proxies_mercado"]],
        start_row=next_row + 1,
    )

    ws_energy = wb.create_sheet("Cadenas_Suministro_Energia")
    _write_table(
        ws_energy,
        ["Ticker", "Nombre", "Precio actual", "Variacion (%)"],
        [[e["ticker"], e["nombre"], e["precio_actual"], e["variacion_pct"]] for e in energy],
    )

    ws_sentiment = wb.create_sheet("Sentimiento_Fondos")
    _write_table(
        ws_sentiment,
        ["Ticker", "Nombre", "Valor actual", "Variacion (%)"],
        [[s["ticker"], s["nombre"], s["precio_actual"], s["variacion_pct"]] for s in sentiment],
    )

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"{_slug(producto)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = REPORTS_DIR / filename
    wb.save(path)
    logger.info("Informe Excel multi-pestana generado: %s", path)
    return path
